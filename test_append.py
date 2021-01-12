#ecoding=utf-8
# author:herui
# time:2021/1/12 8:53
# function:

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pandas import DataFrame
import openpyxl
import os


class TestExcel:

    def test_append_info(self):
        # 1.准备写入的数据
        data = pd.read_excel("test.xlsx","sheet1")
        pf = DataFrame(data)
        # 按索引取值
        col = data.columns.values
        new_data = pf.get(col[2])
        # 按列取值
        # new_data = pf.iloc[0:2,1]
        print()
        print(new_data)
        # n_row = pf.shape[0] # 行数
        # n_col = pf.shape[1] #列数
        # print(f"行数：{n_row},列数：{n_col}")
        # print(data.index)


        # 2.准备待写入的文件
        # 创建新文件：如果存在，则指定写入的sheet；如果不存在，则创建文件和sheet
        if os.path.exists("new_create.xlsx"):
            ws = openpyxl.load_workbook("new_create.xlsx")
            ws.get_sheet_by_name("测试数据")
        else:
            ws = Workbook()
            ws.create_sheet("测试数据")
            ws.save("new_create2.xlsx")
        new_file = ws.active
        n_col = new_file.max_column + 1
        n_row = 2
        # 添加表头
        new_file.cell(1,n_col,col[2])
        for value in new_data:
            new_file.cell(n_row,n_col,value)
            n_row = n_row +1
            # print(n_row,n_col)
        ws.save("new_create.xlsx")

    def test_equal_or_not(self):
        data = pd.read_excel("new_create.xlsx", "测试数据")
        df = DataFrame(data)

        # 按列取值,取第一列和第2列，后面进行比较
        col_1 = df.iloc[:,0]
        col_2 = df.iloc[:,1]

        num = max(len(col_1),len(col_2))
        ws = openpyxl.load_workbook("new_create.xlsx")
        book_sheet = ws.get_sheet_by_name("测试数据")
        # booksheet = ws.worksheets[0]
        book_sheet.insert_cols(3)
        new_col =  ws.active
        row_n = 2

        #设置红色加粗,不相等的时候为红色，相等为绿色
        red_fill = PatternFill("solid", fgColor="FF0000")
        green_fill = PatternFill("solid", fgColor="008000")

        for i in range(0,num):
            if col_1[i] == col_2[i]:
                new_col.cell(row_n,3,"相等").fill=green_fill
            else:
                new_col.cell(row_n, 3, "不一致").fill=red_fill
            row_n = row_n + 1
        ws.save("new_create.xlsx")






